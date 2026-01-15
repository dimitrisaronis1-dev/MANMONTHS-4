import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font
from datetime import datetime
from dateutil.relativedelta import relativedelta
import re
import io
import random

# --- Helper Functions (Copied from your notebook) ---
def parse_date(text, is_start=True):
    text = str(text).strip()
    if re.match(r"^\d{4}$", text):
        if is_start:
            return datetime.strptime("01/01/" + text, "%d/%m/%Y")
        else:
            return datetime.strptime("31/12/" + text, "%d/%m/%Y")
    elif re.match(r"^\d{2}/\d{4}$", text):
        if is_start:
            return datetime.strptime("01/" + text, "%d/%m/%Y")
        else:
            d = datetime.strptime("01/" + text, "%d/%m/%Y")
            return d + relativedelta(months=1) - relativedelta(days=1)
    elif re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", text):
        return datetime.strptime(text, "%d/%m/%Y")
    else:
        raise ValueError(f"Unsupported date format: '{text}'. Expected 'YYYY', 'MM/YYYY', or 'DD/MM/YYYY'.")

def parse_period(p):
    p_cleaned = str(p).strip().replace("—", "-").replace("–", "-")
    if re.match(r"^\d{4}$", p_cleaned):
        return parse_date(p_cleaned, True), parse_date(p_cleaned, False)
    parts = p_cleaned.split("-")
    if len(parts) != 2:
        raise ValueError(f"Invalid period format: '{p}'. Expected 'YYYY' or 'START_DATE-END_DATE'.")
    a, b = parts
    return parse_date(a, True), parse_date(b, False)

def month_range(start, end):
    current = datetime(start.year, start.month, 1)
    end = datetime(end.year, end.month, 1)
    out = []
    while current <= end:
        out.append((current.year, current.month))
        current += relativedelta(months=1)
    return out

def is_light_color(hex_color):
    hex_color = hex_color.lstrip('#')
    rgb = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
    luminance = (0.299 * rgb[0] + 0.587 * rgb[1] + 0.114 * rgb[2]) / 255
    return luminance > 0.5


def process_excel_data(input_file_buffer, template_file_buffer):
    # Define styles
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Read input Excel
    wb_in = openpyxl.load_workbook(input_file_buffer)
    ws_in = wb_in.active

    headers = {}
    for c in range(1, ws_in.max_column + 1):
        val = str(ws_in.cell(1, c).value).strip()
        headers[val] = c

    if "ΧΡΟΝΙΚΟ ΔΙΑΣΤΗΜΑ" not in headers or "ΑΝΘΡΩΠΟΜΗΝΕΣ" not in headers:
        raise Exception("Το input πρέπει να έχει στήλες: ΧΡΟΝΙΚΟ ΔΙΑΣΤΗΜΑ και ΑΝΘΡΩΠΟΜΗΝΕΣ")

    PERIOD_COL = headers["ΧΡΟΝΙΚΟ ΔΙΑΣΤΗΜΑ"]
    AM_COL = headers["ΑΝΘΡΩΠΟΜΗΝΕΣ"]

    data = []
    all_months = set()
    project_counter = 0

    for r in range(2, ws_in.max_row + 1):
        period = ws_in.cell(r, PERIOD_COL).value
        am_raw = ws_in.cell(r, AM_COL).value
        try:
            am = int(am_raw) if am_raw is not None else 0
        except (ValueError, TypeError):
            am = 0

        if not period or am == 0:
            continue

        try:
            start, end = parse_period(str(period))
        except ValueError as e:
            st.warning(f"Skipping row {r} due to period parsing error: {e}")
            continue

        months = month_range(start, end)
        months_in_period_count = len(months)

        if months_in_period_count > 0:
            am_per_month_ratio = am / months_in_period_count
        else:
            am_per_month_ratio = 0

        data.append({
            "project_id": project_counter,
            "period_str": period,
            "original_am": am,
            "months_in_period": months,
            "months_in_period_count": months_in_period_count,
            "am_per_month_ratio": am_per_month_ratio,
            "allocated_am": 0,
            "unallocated_am": am
        })
        project_counter += 1

        for m_val in months:
            all_months.add(m_val)

    all_months = sorted(list(all_months))
    years = sorted(list(set(y for y, m in all_months)))

    data.sort(key=lambda x: x["months_in_period_count"])

    # Open template Excel
    wb = openpyxl.load_workbook(template_file_buffer)
    ws = wb.active

    START_ROW_DATA = 4
    YEAR_ROW = 2
    MONTH_ROW = 3
    YEARLY_TOTAL_ROW = START_ROW_DATA + 1
    START_COL = 5

    # Clear old data
    merged_cells_to_unmerge = []
    for cell_range_str in list(ws.merged_cells.ranges):
        min_col_mc, min_row_mc, max_col_mc, max_row_mc = openpyxl.utils.cell.range_boundaries(str(cell_range_str))
        if ((min_row_mc <= YEAR_ROW <= max_row_mc) or
            (min_row_mc <= MONTH_ROW <= max_col_mc) or
            (min_row_mc <= START_ROW_DATA <= max_row_mc) or
            (min_row_mc <= YEARLY_TOTAL_ROW <= max_row_mc)):
            merged_cells_to_unmerge.append(cell_range_str)

    for cell_range_str in merged_cells_to_unmerge:
        ws.unmerge_cells(str(cell_range_str))

    max_col_to_clear = max(START_COL + len(years) * 12, ws.max_column + 1)
    rows_to_clear_completely = [YEAR_ROW, MONTH_ROW, START_ROW_DATA, YEARLY_TOTAL_ROW]

    for r_clear in rows_to_clear_completely:
        for c_clear in range(1, max_col_to_clear):
            ws.cell(r_clear, c_clear).value = None
            ws.cell(r_clear, c_clear).fill = PatternFill()

    for r_clear in range(START_ROW_DATA + 2, ws.max_row + 1):
        for c_clear in range(1, max_col_to_clear):
            ws.cell(r_clear, c_clear).value = None
            ws.cell(r_clear, c_clear).fill = PatternFill()

    yearly_am_totals = {year: 0 for year in years}
    month_allocation_status = {(y, m): None for y in years for m in range(1, 13)}

    col = START_COL
    month_col_map = {}

    for y in years:
        year_start_col = col
        year_header_cell = ws.cell(YEAR_ROW, col)

        r_color_func = lambda: random.randint(0, 255)
        random_color_hex = '%02X%02X%02X' % (r_color_func(), r_color_func(), r_color_func())
        year_header_cell.fill = PatternFill(start_color=random_color_hex, end_color=random_color_hex, fill_type="solid")

        if not is_light_color(random_color_hex):
            year_header_cell.font = Font(color="FFFFFF")
        else:
            year_header_cell.font = Font(color="000000")

        for m_val in range(1, 13):
            ws.cell(MONTH_ROW, col).value = m_val
            month_col_map[(y, m_val)] = col
            col += 1
        year_end_col = col - 1

        ws.merge_cells(start_row=YEAR_ROW, start_column=year_start_col, end_row=YEAR_ROW, end_column=year_end_col)
        year_header_cell.value = y

    for c_border in range(START_COL, col):
        ws.cell(YEAR_ROW, c_border).border = thin_border
        ws.cell(MONTH_ROW, c_border).border = thin_border

    ws.cell(YEARLY_TOTAL_ROW, 2).value = "ΕΤΗΣΙΑ ΣΥΝΟΛΑ"
    ws.cell(YEARLY_TOTAL_ROW, 2).font = Font(bold=True)
    ws.cell(YEARLY_TOTAL_ROW, 2).border = thin_border

    row = START_ROW_DATA + 2
    unallocated_projects = []
    yearly_overages = {}

    MAX_YEARLY_CAPACITY = 11 # From your notebook's state

    for project_idx, project_data in enumerate(data):
        period_str = project_data["period_str"]
        original_am = project_data["original_am"]
        months_in_period = project_data["months_in_period"]
        project_id = project_data["project_id"]
        allocated_count = 0
        unallocated_count = original_am
        project_unallocated_reason = []

        ws.cell(row, 2).value = period_str
        ws.cell(row, 2).border = thin_border
        ws.cell(row, 3).value = original_am
        ws.cell(row, 3).border = thin_border

        for (y, m_val) in sorted(months_in_period):
            if allocated_count >= original_am:
                break

            if (y, m_val) in month_col_map:
                if yearly_am_totals[y] >= MAX_YEARLY_CAPACITY:
                    if f"Year {y} capacity reached" not in project_unallocated_reason:
                        project_unallocated_reason.append(f"Year {y} capacity reached")
                    continue

                if month_allocation_status[(y, m_val)] is not None:
                    occupying_project_id = month_allocation_status[(y, m_val)]
                    if f"Month {m_val}/{y} already allocated by Project {occupying_project_id}" not in project_unallocated_reason:
                        project_unallocated_reason.append(f"Month {m_val}/{y} already allocated by Project {occupying_project_id}")
                    continue

                cell_to_fill = ws.cell(row, month_col_map[(y, m_val)])
                cell_to_fill.value = 'X'
                cell_to_fill.fill = yellow
                cell_to_fill.border = thin_border

                yearly_am_totals[y] += 1
                month_allocation_status[(y, m_val)] = project_id
                allocated_count += 1
                unallocated_count -= 1

        project_data["allocated_am"] = allocated_count
        project_data["unallocated_am"] = unallocated_count

        if unallocated_count > 0:
            unallocated_projects.append({
                "period": period_str,
                "original_am": original_am,
                "allocated_am": allocated_count,
                "unallocated_am": unallocated_count,
                "reasons": "; ".join(list(set(project_unallocated_reason)))
            })
            ws.cell(row, 3).font = Font(color="FF0000", bold=True)
        else:
            ws.cell(row, 3).font = Font(color="000000")

        for c_border in range(START_COL, col):
            ws.cell(row, c_border).border = thin_border

        row += 1

    for y in years:
        if y in yearly_am_totals:
            col_for_year_total = month_col_map.get((y, 1)) # Get first month's column for the year
            if col_for_year_total is None: # Handle cases where a year has no months in month_col_map
                 continue

            year_month_cols = [month_col_map[(y, m_val)] for m_val in range(1, 13) if (y, m_val) in month_col_map]
            if year_month_cols:
                year_start_col_total = min(year_month_cols)
                year_end_col_total = max(year_month_cols)
                if year_start_col_total != year_end_col_total:
                    ws.merge_cells(start_row=YEARLY_TOTAL_ROW, start_column=year_start_col_total, end_row=YEARLY_TOTAL_ROW, end_column=year_end_col_total)
                col_for_year_total = year_start_col_total
            else:
                continue

            total_cell = ws.cell(YEARLY_TOTAL_ROW, col_for_year_total)
            total_cell.value = yearly_am_totals[y]
            total_cell.font = Font(bold=True)
            total_cell.border = thin_border
            total_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            if yearly_am_totals[y] >= MAX_YEARLY_CAPACITY:
                year_header_cell = ws.cell(YEAR_ROW, col_for_year_total)
                year_header_cell.fill = red_fill
                year_header_cell.font = Font(color="FFFFFF", bold=True)

                total_cell.fill = red_fill
                total_cell.font = Font(color="FFFFFF", bold=True)
                if yearly_am_totals[y] > MAX_YEARLY_CAPACITY:
                    yearly_overages[y] = yearly_am_totals[y] - MAX_YEARLY_CAPACITY
            elif yearly_am_totals[y] > 0:
                total_cell.fill = green_fill
                total_cell.font = Font(color="000000", bold=True)

    for c_width in range(START_COL, col):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c_width)].width = 2.5

    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)

    return output_buffer, yearly_am_totals, unallocated_projects, yearly_overages, MAX_YEARLY_CAPACITY


st.set_page_config(layout="wide")
st.title("Εφαρμογή Κατανομής Ανθρωπομηνών")

st.write("Παρακαλώ ανεβάστε το αρχείο template (οποιοδήποτε excel) και το αρχείο input (με 2 στήλες: 'ΧΡΟΝΙΚΟ ΔΙΑΣΤΗΜΑ' και 'ΑΝΘΡΩΠΟΜΗΝΕΣ').")

template_file = st.file_uploader("Ανεβάστε το αρχείο template (Excel)", type=["xlsx"])
input_file = st.file_uploader("Ανεβάστε το αρχείο input (Excel)", type=["xlsx"])

if template_file is not None and input_file is not None:
    st.success("Και τα δύο αρχεία ανέβηκαν επιτυχώς!")

    # Process the files
    if st.button("Εκτέλεση Κατανομής"):
        try:
            with st.spinner('Επεξεργασία δεδομένων...'):
                output_excel_buffer, yearly_am_totals, unallocated_projects, yearly_overages, MAX_YEARLY_CAPACITY = process_excel_data(input_file, template_file)

            st.subheader("--- Περίληψη Κατανομής ---")
            st.write(f"Μέγιστη ετήσια χωρητικότητα ανά έτος: {MAX_YEARLY_CAPACITY} ανθρωπομήνες")

            st.subheader("Ετήσια Σύνολα Ανθρωπομηνών:")
            for year, total_am in sorted(yearly_am_totals.items()):
                status = "" # Initialize status for each year
                if year in yearly_overages:
                    status = f" (ΥΠΕΡΒΑΣΗ ΧΩΡΗΤΙΚΟΤΗΤΑΣ κατά {yearly_overages[year]})"
                elif total_am >= MAX_YEARLY_CAPACITY:
                    status = " (Η χωρητικότητα έφτασε)"
                st.write(f"  Έτος {year}: {total_am}{status}")

            if unallocated_projects:
                st.subheader("Έργα με μη κατανεμημένους ανθρωπομήνες:")
                for proj in unallocated_projects:
                    st.write(f"  Περίοδος: {proj['period']}, Αρχικοί ΑΜ: {proj['original_am']}, Κατανεμημένοι ΑΜ: {proj['allocated_am']}, Μη κατανεμημένοι ΑΜ: {proj['unallocated_am']}")
                    if proj['reasons']:
                        st.write(f"    Λόγοι για μη κατανομή: {proj['reasons']}")
            else:
                st.write("\nΌλοι οι ανθρωπομήνες κατανεμήθηκαν επιτυχώς.")

            st.download_button(
                label="Κατέβασε το Αρχείο Output",
                data=output_excel_buffer,
                file_name="OUTPUT.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"Προέκυψε σφάλμα κατά την επεξεργασία: {e}")
else:
    st.info("Παρακαλώ ανεβάστε τα αρχεία για να ξεκινήσετε την επεξεργασία.")